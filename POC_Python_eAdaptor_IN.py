from POC_Python_eAdaptor import CargoWise_eAdaptor
from openpyxl import Workbook

class CargoWise_eAdaptor_IN_CargoItineraryData():
    
    def __init__(self, eAdaptor):
        self.eAdaptor = eAdaptor
        
    def main(self):
        IN_eAdaptor = self.eAdaptor
        filename = f"{self.eAdaptor.comp_code}_{self.eAdaptor.key}_CargoItineraryData.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "CargoItinerary Data"
        ws['A1'] = "SR No."
        ws['B1']= "BL No."
        ws['C1'] = "Sequence"
        ws['D1'] = "Port Call Code"
        ws['E1'] = "Next Port Code"
        ws['F1'] = "Mode of Transport"

        mode_map = {
            "SEA": "1",
            "RAI": "2",
            "ROA": "3",
            "AIR": "4"
        }

        transport_mode = IN_eAdaptor.get_text(".//cw:Shipment/cw:TransportMode/cw:Code")
        mode_value = mode_map.get(transport_mode, "")

        ws.append([
            "",
            IN_eAdaptor.get_text(".//cw:Shipment//cw:WayBillNumber"),
            "",
            IN_eAdaptor.get_text(".//cw:Shipment/cw:PortOfLoading/cw:Code"),
            IN_eAdaptor.get_text(".//cw:Shipment/cw:PortOfDischarge/cw:Code"),
            mode_value

        ])


        wb.save(filename)


class CargoWise_eAdaptor_IN_CargoItemsData():
    
    def __init__(self, eAdaptor):
        self.eAdaptor = eAdaptor
        
    def main(self):
        IN_eAdaptor = self.eAdaptor
        filename = f"{self.eAdaptor.comp_code}_{self.eAdaptor.key}_CargoItemsData.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "CargoItems Data"
        ws['A1'] = "SR No."
        ws['B1']= "BL No."
        ws['C1'] = "HS Code"
        ws['D1'] = "CargoItems Description"
        ws['E1'] = "IMO Code"
        ws['F1'] = "UNO Code"
        ws['G1'] = "No. Of Packages"
        ws['H1'] = "Type Of Packages"

        ws.append([
            "",
            IN_eAdaptor.get_text(".//cw:Shipment//cw:WayBillNumber"),
            IN_eAdaptor.get_text(".//cw:Shipment//cw:SubShipmentCollection/cw:SubShipment/cw:PackingLineCollection/cw:PackingLine/cw:HarmonisedCode"),
            IN_eAdaptor.get_text(".//cw:Shipment//cw:SubShipmentCollection/cw:SubShipment/cw:PackingLineCollection/cw:PackingLine/cw:DetailedGoodsDescription"),
            "",
            "",
            IN_eAdaptor.get_text(".//cw:Shipment//cw:SubShipmentCollection/cw:SubShipment/cw:PackingLineCollection/cw:PackingLine/cw:PackQty"),
            IN_eAdaptor.get_text(".//cw:Shipment//cw:SubShipmentCollection/cw:SubShipment/cw:PackingLineCollection/cw:PackingLine/cw:PackType/cw:Code")           

        ])


        wb.save(filename)

class CargoWise_eAdaptor_IN_ContainerData():
    
    def __init__(self, eAdaptor):
        self.eAdaptor = eAdaptor
        
    def main(self):
        IN_eAdaptor = self.eAdaptor
        filename = f"{self.eAdaptor.comp_code}_{self.eAdaptor.key}_ContainerData.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Container Data"

        ws['A1'] = "SR No."
        ws['B1']= "BL No."
        ws['C1'] = "Continer No."
        ws['D1'] = "ISO Code"
        ws['E1'] = "Container Status"
        ws['F1'] = "Add Eqp Hold"
        ws['G1'] = "Seal Type"
        ws['H1'] = "Seal No 1"
        ws['I1'] = "Seal No 2"
        ws['J1'] = "Seal No 3"
        ws['K1'] = "Other Equipment ID"
        ws['L1'] = "SOC Flag"
        ws['M1'] = "Agent Code"
        ws['N1'] = "Container Weight"
        ws['O1'] = "No. Of Packages"




        ws.append([
            "",
            IN_eAdaptor.get_text(".//cw:Shipment//cw:WayBillNumber"),
            ", ".join([c.text for c in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:ContainerNumber", IN_eAdaptor.ns) if c.text]),
            ", ".join([c.text for c in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:ContainerType/cw:Code", IN_eAdaptor.ns) if c.text]),
            ", ".join([c.text for c in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:ContainerStatus/cw:Code", IN_eAdaptor.ns) if c.text]),
            "",
            "",
            ", ".join([s1.text for s1 in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:Seal1", IN_eAdaptor.ns) if s1.text]),
            ", ".join([s2.text for s2 in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:Seal2", IN_eAdaptor.ns) if s2.text]),
            ", ".join([s3.text for s3 in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:Seal3", IN_eAdaptor.ns) if s3.text]),
            "",
            "",
            ", ".join([c.text for c in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:IsShipperOwned",IN_eAdaptor.ns) if c.text]),
            ", ".join([c.text for c in IN_eAdaptor.root.findall(".//cw:Shipment/cw:ContainerCollection/cw:Container/cw:GrosWeight", IN_eAdaptor.ns) if c.text]),
            ""
        ])


        wb.save(filename)
